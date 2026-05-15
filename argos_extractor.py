#!/usr/bin/env python3
"""
ARGOS Extractor → SQL para Supabase
====================================
Uso:
    python argos_extractor.py IARD.xlsx [VIDA.xlsx]

Genera sentencias SQL listas para pegar en el Editor SQL de Supabase.
El COR se toma directamente de la columna COR (0 = dato aún no disponible).
"""

import sys
import json
import re
import openpyxl
from pathlib import Path

# ─── Constantes ───────────────────────────────────────────────────────────────

MEDOFI_CODES = {'742776', '755224', '742826', '742821', '742827', '742825'}

MONTH_NAMES_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
}

# ─── Utilidades ───────────────────────────────────────────────────────────────

def parse_month_year(text: str):
    """Extrae mes y año del nombre del fichero."""
    text_upper = text.upper()
    month = next((n for m, n in MONTH_NAMES_ES.items() if m in text_upper), None)
    m = re.search(r'20\d{2}', text)
    year = int(m.group()) if m else None
    return month, year


def get_col_index(ws) -> dict:
    """Encuentra la fila de cabeceras y devuelve un dict {nombre: índice}."""
    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        first = str(row[0]) if row[0] else ''
        if first in ('LoB', 'Negocio'):
            return {str(h).strip(): i for i, h in enumerate(row) if h}
    return {}


def fget(row, col: dict, name: str, pct: bool = False) -> float:
    """Lee un valor de la fila de forma segura. Si pct=True multiplica x100."""
    idx = col.get(name)
    if idx is None or idx >= len(row):
        return 0.0
    v = row[idx]
    if v is None:
        return 0.0
    try:
        result = float(v)
        return round(result * 100, 4) if pct else round(result, 4)
    except (TypeError, ValueError):
        return 0.0


def classify_iard_sheet(name: str) -> str | None:
    """
    Devuelve:
      'GLOBAL'   → hoja medor / nombre de mes
      'COR'      → hoja con datos de COR
      código     → medofi concreto (ej: '742776')
      None       → ignorar
    """
    upper = name.upper()

    if 'COR' in upper:
        return 'COR'

    # Medor explícito (ej: "742776 medor")
    if 'MEDOR' in upper and 'MEDOFI' not in upper:
        return 'GLOBAL'

    # Medofi con código (ej: "742776 medofi")
    if 'MEDOFI' in upper:
        for code in MEDOFI_CODES:
            if code in name:
                return code
        return None

    # Nombre es exactamente un código medofi
    if name.strip() in MEDOFI_CODES:
        return name.strip()

    # Hoja cuyo nombre es un mes (ej: "ABRIL", "SEPTIEMBRE")
    for m in MONTH_NAMES_ES:
        if m in upper:
            return 'GLOBAL'

    return None


# ─── Extracción IARD ──────────────────────────────────────────────────────────

def empty_record() -> dict:
    return {
        'medofis': {
            'gwp': 0, 'crecimientoPct': 0, 'renovacionPct': 0,
            'tasaNpPct': 0, 'cor': 0, 'devolucionesPct': 0,
            'siniestralidadSinIbnrPct': 0
        },
        'cartera': {
            'particulares': {'auto': 0, 'hogar': 0, 'comunidades': 0,
                             'decesos': 0, 'rc': 0, 'salud': 0, 'total': 0},
            'empresa':      {'rc': 0, 'flotas': 0, 'comercio': 0,
                             'oficina': 0, 'industria': 0, 'transporte': 0, 'total': 0},
            'vida':         {'individual': 0, 'ahorro': 0},
            'psc':          {'vida': 0, 'salud': 0, 'total': 0}
        },
        'produccion': {
            'particulares': {'auto': 0, 'hogar': 0, 'comunidades': 0,
                             'decesos': 0, 'rc': 0, 'salud': 0, 'total': 0},
            'empresa':      {'rc': 0, 'flotas': 0, 'comercio': 0,
                             'oficina': 0, 'industria': 0, 'transporte': 0, 'total': 0},
            'vida':         {'individual': 0, 'ahorro': 0},
            'psc':          {'vida': 0, 'salud': 0, 'total': 0}
        }
    }


def extract_iard(ws) -> dict | None:
    col = get_col_index(ws)
    if not col:
        return None

    r = empty_record()
    current_lob   = None
    current_agrup = None

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        v0 = row[0]
        v1 = row[1] if len(row) > 1 else None
        v2 = row[2] if len(row) > 2 else None

        # Ignorar cabeceras y notas de filtros
        if v0 in ('LoB', 'Negocio'):
            continue
        if v0 and 'Filtros' in str(v0):
            continue

        # Actualizar contexto
        if v0 and v0 != 'Total':
            current_lob = str(v0).upper()
        if v1 and v1 != 'Total':
            current_agrup = str(v1).upper()

        # ── TOTAL GENERAL (v0='Total', v1=None) ──────────────────────────────
        if v0 == 'Total' and v1 is None:
            r['medofis'] = {
                'gwp':                    fget(row, col, 'GWP'),
                'crecimientoPct':         fget(row, col, '% GWP', pct=True),
                'renovacionPct':          fget(row, col, '% Renov. PRIMAS', pct=True),
                'tasaNpPct':              fget(row, col, '% Tasa NP', pct=True),
                'cor':                    fget(row, col, 'COR'),          # 0 = dato no disponible aún
                'devolucionesPct':        fget(row, col, '%PTE P.Adq', pct=True),
                'siniestralidadSinIbnrPct': fget(row, col, 'Siniestralidad sin_IBNR', pct=True),
            }
            continue

        # ── TOTAL POR LOB (v0=None, v1='Total', v2=None) ─────────────────────
        if v0 is None and v1 == 'Total' and v2 is None and current_lob:
            gwp   = fget(row, col, 'GWP')
            gwpnp = fget(row, col, 'GWPNP')
            if 'PART' in current_lob:
                r['cartera']['particulares']['total']   = gwp
                r['produccion']['particulares']['total'] = gwpnp
            elif 'EMP' in current_lob:
                r['cartera']['empresa']['total']   = gwp
                r['produccion']['empresa']['total'] = gwpnp
            elif 'SALUD' in current_lob:
                r['cartera']['particulares']['salud']   = gwp
                r['produccion']['particulares']['salud'] = gwpnp
            continue

        # ── TOTAL POR AGRUPACIÓN FLASH (v0=None, v1=None, v2='Total') ────────
        if v0 is None and v1 is None and v2 == 'Total' and current_agrup:
            gwp   = fget(row, col, 'GWP')
            gwpnp = fget(row, col, 'GWPNP')
            ag = current_agrup

            if 'AUTO' in ag:
                r['cartera']['particulares']['auto']    = gwp
                r['produccion']['particulares']['auto'] = gwpnp
            elif 'HOGAR' in ag:
                r['cartera']['particulares']['hogar']    = gwp
                r['produccion']['particulares']['hogar'] = gwpnp
            elif 'FLOTA' in ag:
                r['cartera']['empresa']['flotas']    = gwp
                r['produccion']['empresa']['flotas'] = gwpnp
            elif 'INDUSTRIA' in ag or 'DA' in ag:
                r['cartera']['empresa']['industria']    = gwp
                r['produccion']['empresa']['industria'] = gwpnp
            elif 'RC EMP' in ag or ('RC' in ag and current_lob and 'EMP' in current_lob):
                r['cartera']['empresa']['rc']    = gwp
                r['produccion']['empresa']['rc'] = gwpnp
            continue

    # ── RESTO: calcular por diferencia ──────────────────────────────────────
    # Particulares: lo que no es auto, hogar ni salud
    c_part = r['cartera']['particulares']
    p_part = r['produccion']['particulares']
    resto_c = c_part['total'] - c_part['auto'] - c_part['hogar'] - c_part['salud']
    resto_p = p_part['total'] - p_part['auto'] - p_part['hogar'] - p_part['salud']
    c_part['comunidades'] = round(max(0.0, resto_c), 2)
    p_part['comunidades'] = round(max(0.0, resto_p), 2)

    # Empresa: lo que no es flotas, industria ni rc
    c_emp = r['cartera']['empresa']
    p_emp = r['produccion']['empresa']
    resto_c_emp = c_emp['total'] - c_emp['flotas'] - c_emp['industria'] - c_emp['rc']
    resto_p_emp = p_emp['total'] - p_emp['flotas'] - p_emp['industria'] - p_emp['rc']
    c_emp['comercio'] = round(max(0.0, resto_c_emp), 2)
    p_emp['comercio'] = round(max(0.0, resto_p_emp), 2)

    return r


def extract_cor_from_sheet(ws) -> float:
    """Extrae el COR del total general de una hoja COR."""
    col = get_col_index(ws)
    if not col:
        return 0.0
    for row in ws.iter_rows(values_only=True):
        v0 = row[0]
        v1 = row[1] if len(row) > 1 else None
        if v0 == 'Total' and v1 is None:
            return fget(row, col, 'COR')
    return 0.0


# ─── Extracción VIDA ──────────────────────────────────────────────────────────

def extract_vida(ws) -> dict | None:
    col = get_col_index(ws)
    if not col:
        return None

    result = {
        'ind_gwp': 0.0, 'ind_gwpnp': 0.0,
        'ahorro_gwp': 0.0, 'ahorro_gwpnp': 0.0,
        'col_gwp': 0.0, 'col_gwpnp': 0.0,
        'renovacionPct': 0.0, 'tasaNpPct': 0.0, 'devolucionesPct': 0.0
    }

    current_negocio = None

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        v0 = row[0]
        v1 = row[1] if len(row) > 1 else None

        if v0 == 'Negocio':
            continue
        if v0 and 'Filtros' in str(v0):
            continue

        if v0 and v0 != 'Total':
            current_negocio = str(v0).upper()

        # Total general de VIDA
        if v0 == 'Total' and v1 is None:
            result['renovacionPct']   = fget(row, col, '% Renov PRIMAS', pct=True)
            result['tasaNpPct']       = fget(row, col, 'Tasa NP', pct=True)
            result['devolucionesPct'] = fget(row, col, '% PTE', pct=True)
            continue

        # Total Individual
        if v0 is None and v1 == 'Total' and current_negocio and 'INDIVIDUAL' in current_negocio:
            result['ind_gwp']   = fget(row, col, 'GWP')
            result['ind_gwpnp'] = fget(row, col, 'GWPNP')
            continue

        # Total Colectivo
        if v0 is None and v1 == 'Total' and current_negocio and 'COLECTIVO' in current_negocio:
            result['col_gwp']   = fget(row, col, 'GWP')
            result['col_gwpnp'] = fget(row, col, 'GWPNP')
            continue

        # Sub-líneas de Individual para separar ahorro/inversión
        if v0 is None and current_negocio and 'INDIVIDUAL' in current_negocio and v1:
            lob_upper = str(v1).upper()
            if any(k in lob_upper for k in ('GENERAL ACCOUNT', 'UNIT LINKED', 'GENERAL', 'LINKED')):
                result['ahorro_gwp']   = round(result['ahorro_gwp']   + fget(row, col, 'GWP'), 2)
                result['ahorro_gwpnp'] = round(result['ahorro_gwpnp'] + fget(row, col, 'GWPNP'), 2)

    # Vida protección = Individual total - Ahorro
    result['vida_gwp']   = round(result['ind_gwp']   - result['ahorro_gwp'], 2)
    result['vida_gwpnp'] = round(result['ind_gwpnp'] - result['ahorro_gwpnp'], 2)

    return result


def merge_vida(record: dict, vida: dict) -> dict:
    """Añade los datos VIDA al registro IARD."""
    record['cartera']['vida']['individual']  = max(0.0, vida['vida_gwp'])
    record['cartera']['vida']['ahorro']      = vida['ahorro_gwp']
    record['cartera']['psc']['vida']         = vida['col_gwp']
    record['cartera']['psc']['total']        = vida['col_gwp']

    record['produccion']['vida']['individual'] = max(0.0, vida['vida_gwpnp'])
    record['produccion']['vida']['ahorro']     = vida['ahorro_gwpnp']
    record['produccion']['psc']['vida']        = vida['col_gwpnp']
    record['produccion']['psc']['total']       = vida['col_gwpnp']

    return record


# ─── Generación SQL ───────────────────────────────────────────────────────────

def to_sql(year: int, month: int, code: str, record: dict) -> str:
    def esc(d):
        return json.dumps(d, ensure_ascii=False).replace("'", "''")

    return (
        f"INSERT INTO metrics (year, month, mediator_code, medofis, cartera, produccion)\n"
        f"VALUES (\n"
        f"  {year}, {month}, '{code}',\n"
        f"  '{esc(record['medofis'])}'::jsonb,\n"
        f"  '{esc(record['cartera'])}'::jsonb,\n"
        f"  '{esc(record['produccion'])}'::jsonb\n"
        f")\n"
        f"ON CONFLICT (year, month, mediator_code)\n"
        f"DO UPDATE SET\n"
        f"  medofis    = EXCLUDED.medofis,\n"
        f"  cartera    = EXCLUDED.cartera,\n"
        f"  produccion = EXCLUDED.produccion;\n"
    )


# ─── Proceso principal ────────────────────────────────────────────────────────

def process(iard_path: str, vida_path: str | None = None):
    iard_file = Path(iard_path)
    month, year = parse_month_year(iard_file.name)

    if not month or not year:
        print("⚠️  No se pudo detectar mes/año del nombre del fichero IARD.")
        print("   Asegúrate de que el nombre incluye el mes en español y el año (ej: SEPTIEMBRE 2025)")
        return

    print(f"-- ================================================================")
    print(f"-- ARGOS → Supabase  |  Mes: {month}  Año: {year}")
    print(f"-- Fichero IARD: {iard_file.name}")
    if vida_path:
        print(f"-- Fichero VIDA: {Path(vida_path).name}")
    print(f"-- Generado automáticamente. Revisa los datos antes de ejecutar.")
    print(f"-- ================================================================\n")

    wb_iard = openpyxl.load_workbook(iard_path, data_only=True)

    # Datos VIDA
    vida_data = None
    if vida_path:
        wb_vida = openpyxl.load_workbook(vida_path, data_only=True)
        for sname in wb_vida.sheetnames:
            vida_data = extract_vida(wb_vida[sname])
            if vida_data:
                break

    # Clasificar y extraer hojas IARD
    records   = {}   # código → record
    global_rec = None
    cor_global = None

    for sname in wb_iard.sheetnames:
        role = classify_iard_sheet(sname)
        if role is None:
            continue

        ws = wb_iard[sname]

        if role == 'COR':
            cor_global = extract_cor_from_sheet(ws)
            continue

        extracted = extract_iard(ws)
        if not extracted:
            continue

        if role == 'GLOBAL':
            global_rec = extracted
        else:
            records[role] = extracted

    # Aplicar COR global si viene de hoja COR separada
    if global_rec and cor_global is not None and global_rec['medofis']['cor'] == 0:
        global_rec['medofis']['cor'] = cor_global

    # Combinar VIDA en el GLOBAL
    if global_rec and vida_data:
        global_rec = merge_vida(global_rec, vida_data)

    # ── Generar SQL ──────────────────────────────────────────────────────────
    total = 0

    if global_rec:
        print(f"-- GLOBAL (MEDOR) --")
        print(to_sql(year, month, 'GLOBAL', global_rec))
        total += 1

    for code in sorted(records):
        label = {
            '742776': 'Badajoz',
            '742821': 'Copttraba',
            '742825': 'Madrid',
            '742826': 'Herrera',
            '742827': 'Olivenza',
            '755224': 'San Roque',
        }.get(code, code)
        print(f"-- Medofi {code} ({label}) --")
        print(to_sql(year, month, code, records[code]))
        total += 1

    # ── Resumen ──────────────────────────────────────────────────────────────
    print(f"-- ================================================================")
    print(f"-- ✅ {total} registro(s) generado(s): "
          f"{'1 GLOBAL' if global_rec else 'sin GLOBAL'} + {len(records)} medofi(s)")
    if not vida_data:
        print(f"-- ⚠️  Sin fichero VIDA → cartera/producción de Vida, Ahorro y PSC = 0")
    if global_rec and global_rec['medofis']['cor'] == 0:
        print(f"-- ℹ️  COR = 0 (dato aún no disponible para este mes)")
    print(f"-- ================================================================")


# ─── Punto de entrada ─────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    iard  = sys.argv[1]
    vida  = sys.argv[2] if len(sys.argv) > 2 else None
    process(iard, vida)
